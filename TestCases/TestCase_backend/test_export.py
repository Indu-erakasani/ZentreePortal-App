"""
Run from project root:
    cd /home/indhu/zentreeportal
    pytest TestCases/TestCase_backend/test_export.py -v
"""
import pytest
import csv
import io
from bson import ObjectId
from datetime import datetime


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _register_user(client, email, role="admin", password="Test@1234"):
    client.post("/api/auth/register", json={
        "first_name": "Test",
        "last_name":  "User",
        "email":      email,
        "password":   password,
        "role":       role,
    })


def _login(client, email, password="Test@1234"):
    res  = client.post("/api/auth/login", json={"email": email, "password": password})
    data = res.get_json()
    return {"Authorization": f"Bearer {data['access_token']}"}


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def auth_headers(client):
    _register_user(client, "export_admin@test.com", role="admin")
    return _login(client, "export_admin@test.com")


@pytest.fixture(scope="module")
def recruiter_headers(client):
    _register_user(client, "export_recruiter@test.com", role="recruiter")
    return _login(client, "export_recruiter@test.com")


@pytest.fixture(scope="module")
def seeded_employee(app):
    """
    Insert one employee + one onboarding record.
    Yields (emp_id_str, mongo_oid_str).
    Cleans up after the module.
    """
    from extensions import mongo
    with app.app_context():
        emp_oid = mongo.db.employees.insert_one({
            "emp_id":              "EXP_EMP_001",
            "name":                "Export Tester",
            "email":               "export.tester@acme.com",
            "phone":               "9876543210",
            "designation":         "QA Engineer",
            "department":          "Engineering",
            "employment_type":     "Full-Time",
            "date_of_joining":     "2023-01-15T00:00:00",
            "experience":          3,
            "location":            "Hyderabad",
            "reporting_manager":   "Manager One",
            "status":              "Active",
            "skills":              "Python, Selenium",
            "current_client":      "Acme Corp",
            "current_project":     "Portal",
            "current_billing_rate": 5000,
            "billing_currency":    "INR",
            "salary":              800000,
            "notes":               "Test employee",
            "client_history": [
                {
                    "client_name":      "Old Corp",
                    "project_name":     "Legacy App",
                    "role":             "Developer",
                    "start_date":       "2021-06-01",
                    "end_date":         "2022-12-31",
                    "billing_rate":     4000,
                    "billing_currency": "INR",
                    "work_location":    "Remote",
                    "technology":       "Python",
                }
            ],
        }).inserted_id

        mongo.db.onboarding.insert_one({
            "employee_id":    str(emp_oid),
            "blood_group":    "O+",
            "personal_email": "tester.personal@gmail.com",
            "referred_by":    "HR Team",
            "bgv_status":     "Completed",
            "bgv_agency":     "Verifact",
            "bgv_remarks":    "Clear",
            "laptop_serial":  "SN12345",
            "laptop_make_model": "Dell XPS 15",
            "access_card_number": "AC-001",
            "email_id_created":   "export.tester@company.com",
            "probation_end_date": "2023-04-15",
            "bank_details": {
                "account_holder_name": "Export Tester",
                "account_number":      "1234567890",
                "ifsc_code":           "HDFC0001234",
                "bank_name":           "HDFC Bank",
                "branch":              "Hyderabad",
                "account_type":        "Savings",
            },
            "emergency_contact": {
                "name":         "Emergency Person",
                "relationship": "Spouse",
                "phone":        "9000000001",
                "email":        "emergency@gmail.com",
                "address":      "Hyderabad",
            },
            "checklist": [
                {"label": "ID Proof", "done": True,  "remarks": "Verified", "updated_at": "2023-01-16"},
                {"label": "Offer Letter", "done": True,  "remarks": "Signed",    "updated_at": "2023-01-16"},
                {"label": "Laptop Issue", "done": False, "remarks": "",           "updated_at": ""},
            ],
            "documents": [
                {"name": "Aadhar Card", "category": "ID Proof",   "status": "Verified",  "remarks": "OK",      "file_name": "aadhar.pdf",  "updated_at": "2023-01-17"},
                {"name": "PAN Card",    "category": "Tax",         "status": "Pending",   "remarks": "Awaited", "file_name": None,           "updated_at": ""},
            ],
            "hr_notes": "Good candidate",
            "it_notes": "Laptop issued",
        })

    yield "EXP_EMP_001", str(emp_oid)

    # Teardown
    with app.app_context():
        from extensions import mongo as m
        m.db.employees.delete_one({"_id": emp_oid})
        m.db.onboarding.delete_one({"employee_id": str(emp_oid)})


# ═════════════════════════════════════════════════════════════════════════════
# 1.  GET /api/export/employees/csv
# ═════════════════════════════════════════════════════════════════════════════

class TestExportCSV:

    # ── Auth ──────────────────────────────────────────────────────────────────

    def test_unauthenticated_returns_401(self, client):
        assert client.get("/api/export/employees/csv").status_code == 401

    def test_authenticated_returns_200(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=auth_headers)
        assert res.status_code == 200

    def test_recruiter_can_access(self, client, recruiter_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=recruiter_headers)
        assert res.status_code == 200

    # ── Response headers ──────────────────────────────────────────────────────

    def test_content_type_is_csv(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=auth_headers)
        assert "text/csv" in res.content_type

    def test_content_disposition_is_attachment(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=auth_headers)
        assert "attachment" in res.headers.get("Content-Disposition", "")

    def test_filename_contains_date(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=auth_headers)
        cd  = res.headers.get("Content-Disposition", "")
        today = datetime.utcnow().strftime("%Y%m%d")
        assert today in cd

    def test_filename_ends_with_csv(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/csv", headers=auth_headers)
        cd  = res.headers.get("Content-Disposition", "")
        assert ".csv" in cd

    # ── CSV content ───────────────────────────────────────────────────────────

    def test_csv_is_parseable(self, client, auth_headers, seeded_employee):
        res  = client.get("/api/export/employees/csv", headers=auth_headers)
        text = res.data.decode("utf-8")
        rows = list(csv.DictReader(io.StringIO(text)))
        assert isinstance(rows, list)

    def test_csv_has_rows(self, client, auth_headers, seeded_employee):
        res  = client.get("/api/export/employees/csv", headers=auth_headers)
        text = res.data.decode("utf-8")
        rows = list(csv.DictReader(io.StringIO(text)))
        assert len(rows) >= 1

    def test_csv_has_emp_id_header(self, client, auth_headers, seeded_employee):
        res     = client.get("/api/export/employees/csv", headers=auth_headers)
        text    = res.data.decode("utf-8")
        headers = text.splitlines()[0]
        assert "Emp ID" in headers

    def test_csv_has_full_name_header(self, client, auth_headers, seeded_employee):
        res     = client.get("/api/export/employees/csv", headers=auth_headers)
        text    = res.data.decode("utf-8")
        assert "Full Name" in text.splitlines()[0]

    def test_csv_expected_columns_present(self, client, auth_headers, seeded_employee):
        res     = client.get("/api/export/employees/csv", headers=auth_headers)
        text    = res.data.decode("utf-8")
        header  = text.splitlines()[0]
        for col in ("Email", "Phone", "Designation", "Department",
                    "Status", "Date of Joining", "Annual Salary"):
            assert col in header, f"Missing column: {col}"

    def test_csv_bank_columns_present(self, client, auth_headers, seeded_employee):
        res    = client.get("/api/export/employees/csv", headers=auth_headers)
        text   = res.data.decode("utf-8")
        header = text.splitlines()[0]
        for col in ("Bank Name", "IFSC Code", "Bank Account Number"):
            assert col in header, f"Missing column: {col}"

    def test_csv_emergency_contact_columns_present(self, client, auth_headers, seeded_employee):
        res    = client.get("/api/export/employees/csv", headers=auth_headers)
        text   = res.data.decode("utf-8")
        header = text.splitlines()[0]
        assert "Emergency Contact Name" in header
        assert "Emergency Contact Phone" in header

    def test_csv_contains_seeded_employee(self, client, auth_headers, seeded_employee):
        emp_id_str, _ = seeded_employee
        res  = client.get("/api/export/employees/csv", headers=auth_headers)
        text = res.data.decode("utf-8")
        assert "Export Tester" in text

    def test_csv_contains_seeded_emp_id(self, client, auth_headers, seeded_employee):
        emp_id_str, _ = seeded_employee
        res  = client.get("/api/export/employees/csv", headers=auth_headers)
        text = res.data.decode("utf-8")
        assert emp_id_str in text


# ═════════════════════════════════════════════════════════════════════════════
# 2.  GET /api/export/employees/excel
# ═════════════════════════════════════════════════════════════════════════════

class TestExportExcel:

    EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ── Auth ──────────────────────────────────────────────────────────────────

    def test_unauthenticated_returns_401(self, client):
        assert client.get("/api/export/employees/excel").status_code == 401

    def test_authenticated_returns_200(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        assert res.status_code == 200

    # ── Response headers ──────────────────────────────────────────────────────

    def test_content_type_is_xlsx(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        assert self.EXCEL_MIME in res.content_type

    def test_content_disposition_is_attachment(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        assert "attachment" in res.headers.get("Content-Disposition", "")

    def test_filename_ends_with_xlsx(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        cd  = res.headers.get("Content-Disposition", "")
        assert ".xlsx" in cd

    def test_filename_contains_date(self, client, auth_headers, seeded_employee):
        res   = client.get("/api/export/employees/excel", headers=auth_headers)
        cd    = res.headers.get("Content-Disposition", "")
        today = datetime.utcnow().strftime("%Y%m%d")
        assert today in cd

    # ── Workbook content ──────────────────────────────────────────────────────

    def test_response_data_is_non_empty(self, client, auth_headers, seeded_employee):
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        assert len(res.data) > 0

    def test_workbook_is_valid_xlsx(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert wb is not None

    def test_workbook_has_three_sheets(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert len(wb.sheetnames) == 3

    def test_sheet1_is_employee_summary(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Employee Summary" in wb.sheetnames

    def test_sheet2_is_onboarding_checklist(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Onboarding Checklist" in wb.sheetnames

    def test_sheet3_is_documents(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Documents" in wb.sheetnames

    def test_summary_sheet_has_header_row(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        ws  = wb["Employee Summary"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Emp ID" in headers

    def test_summary_sheet_has_data_row(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        ws  = wb["Employee Summary"]
        assert ws.max_row >= 2

    def test_summary_sheet_contains_seeded_employee(self, client, auth_headers, seeded_employee):
        import openpyxl
        res   = client.get("/api/export/employees/excel", headers=auth_headers)
        wb    = openpyxl.load_workbook(io.BytesIO(res.data))
        ws    = wb["Employee Summary"]
        values = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any("Export Tester" in v for v in values)

    def test_checklist_sheet_has_header_row(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        ws  = wb["Onboarding Checklist"]
        assert ws.max_row >= 1

    def test_documents_sheet_has_header_row(self, client, auth_headers, seeded_employee):
        import openpyxl
        res = client.get("/api/export/employees/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        ws  = wb["Documents"]
        assert ws.max_row >= 1


# ═════════════════════════════════════════════════════════════════════════════
# 3.  GET /api/export/employee/<emp_id>  — single employee JSON
# ═════════════════════════════════════════════════════════════════════════════

class TestExportSingleJSON:

    # ── Auth ──────────────────────────────────────────────────────────────────

    def test_unauthenticated_returns_401(self, client, seeded_employee):
        _, oid = seeded_employee
        assert client.get(f"/api/export/employee/{oid}").status_code == 401

    def test_invalid_id_returns_400(self, client, auth_headers):
        assert client.get("/api/export/employee/NOT_AN_OID",
                          headers=auth_headers).status_code == 400

    def test_unknown_id_returns_404(self, client, auth_headers):
        fake_oid = str(ObjectId())
        assert client.get(f"/api/export/employee/{fake_oid}",
                          headers=auth_headers).status_code == 404

    def test_valid_id_returns_200(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        assert client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).status_code == 200

    # ── Response shape ────────────────────────────────────────────────────────

    def test_success_flag_is_true(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        body = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()
        assert body["success"] is True

    def test_data_key_present(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        body = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()
        assert "data" in body

    def test_data_has_emp_id(self, client, auth_headers, seeded_employee):
        emp_id_str, oid = seeded_employee
        data = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()["data"]
        assert data.get("emp_id") == emp_id_str

    def test_data_has_name(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        data = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()["data"]
        assert data.get("name") == "Export Tester"

    def test_data_has_onboarding_key(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        data = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()["data"]
        assert "onboarding" in data

    def test_onboarding_has_bank_details(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        ob = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]
        assert "bank_details" in ob

    def test_onboarding_has_emergency_contact(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        ob = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]
        assert "emergency_contact" in ob

    def test_onboarding_has_checklist(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        ob = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]
        assert "checklist" in ob
        assert isinstance(ob["checklist"], list)

    def test_onboarding_has_documents(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        ob = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]
        assert "documents" in ob
        assert isinstance(ob["documents"], list)

    def test_data_id_is_string(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        data = client.get(f"/api/export/employee/{oid}",
                          headers=auth_headers).get_json()["data"]
        assert isinstance(data.get("_id"), str)

    def test_bank_details_has_ifsc(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        bd = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]["bank_details"]
        assert "ifsc_code" in bd
        assert bd["ifsc_code"] == "HDFC0001234"

    def test_emergency_contact_has_phone(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        ec = client.get(f"/api/export/employee/{oid}",
                        headers=auth_headers).get_json()["data"]["onboarding"]["emergency_contact"]
        assert "phone" in ec


# ═════════════════════════════════════════════════════════════════════════════
# 4.  GET /api/export/employee/<emp_id>/excel  — single employee Excel
# ═════════════════════════════════════════════════════════════════════════════

class TestExportSingleExcel:

    EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ── Auth ──────────────────────────────────────────────────────────────────

    def test_unauthenticated_returns_401(self, client, seeded_employee):
        _, oid = seeded_employee
        assert client.get(f"/api/export/employee/{oid}/excel").status_code == 401

    def test_invalid_id_returns_400(self, client, auth_headers):
        assert client.get("/api/export/employee/BAD_ID/excel",
                          headers=auth_headers).status_code == 400

    def test_unknown_id_returns_404(self, client, auth_headers):
        fake_oid = str(ObjectId())
        assert client.get(f"/api/export/employee/{fake_oid}/excel",
                          headers=auth_headers).status_code == 404

    def test_valid_id_returns_200(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        assert client.get(f"/api/export/employee/{oid}/excel",
                          headers=auth_headers).status_code == 200

    # ── Response headers ──────────────────────────────────────────────────────

    def test_content_type_is_xlsx(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        assert self.EXCEL_MIME in res.content_type

    def test_content_disposition_is_attachment(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        assert "attachment" in res.headers.get("Content-Disposition", "")

    def test_filename_contains_emp_id(self, client, auth_headers, seeded_employee):
        emp_id_str, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        cd  = res.headers.get("Content-Disposition", "")
        assert emp_id_str in cd

    def test_filename_ends_with_xlsx(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        cd  = res.headers.get("Content-Disposition", "")
        assert ".xlsx" in cd

    def test_filename_contains_date(self, client, auth_headers, seeded_employee):
        _, oid = seeded_employee
        res   = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        cd    = res.headers.get("Content-Disposition", "")
        today = datetime.utcnow().strftime("%Y%m%d")
        assert today in cd

    # ── Workbook content ──────────────────────────────────────────────────────

    def test_workbook_is_valid_xlsx(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert wb is not None

    def test_workbook_has_four_sheets(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert len(wb.sheetnames) == 4

    def test_sheet_profile_present(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Profile" in wb.sheetnames

    def test_sheet_onboarding_checklist_present(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Onboarding Checklist" in wb.sheetnames

    def test_sheet_documents_present(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Documents" in wb.sheetnames

    def test_sheet_client_history_present(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb  = openpyxl.load_workbook(io.BytesIO(res.data))
        assert "Client History" in wb.sheetnames

    def test_profile_sheet_contains_employee_name(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res    = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb     = openpyxl.load_workbook(io.BytesIO(res.data))
        ws     = wb["Profile"]
        values = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any("Export Tester" in v for v in values)

    def test_checklist_sheet_has_done_items(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res    = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb     = openpyxl.load_workbook(io.BytesIO(res.data))
        ws     = wb["Onboarding Checklist"]
        values = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any("Done" in v for v in values)

    def test_documents_sheet_has_verified_status(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res    = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb     = openpyxl.load_workbook(io.BytesIO(res.data))
        ws     = wb["Documents"]
        values = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any("Verified" in v for v in values)

    def test_client_history_sheet_has_data(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res    = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb     = openpyxl.load_workbook(io.BytesIO(res.data))
        ws     = wb["Client History"]
        # header row + at least 1 data row
        assert ws.max_row >= 3

    def test_client_history_contains_old_client(self, client, auth_headers, seeded_employee):
        import openpyxl
        _, oid = seeded_employee
        res    = client.get(f"/api/export/employee/{oid}/excel", headers=auth_headers)
        wb     = openpyxl.load_workbook(io.BytesIO(res.data))
        ws     = wb["Client History"]
        values = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any("Old Corp" in v for v in values)


# ═════════════════════════════════════════════════════════════════════════════
# 5.  Unit tests — _flat_row helper
# ═════════════════════════════════════════════════════════════════════════════

class TestFlatRow:

    def _fn(self):
        from zentreeportal_backend.routes.export_routes import _flat_row
        return _flat_row

    @pytest.fixture
    def sample_emp(self):
        return {
            "emp_id": "E001", "name": "Alice", "email": "a@b.com",
            "phone": "9999", "designation": "Dev", "department": "Eng",
            "employment_type": "FT", "date_of_joining": "2023-01-01T00:00:00",
            "experience": 3, "location": "Hyd", "reporting_manager": "Mgr",
            "status": "Active", "skills": "Python", "current_client": "XYZ",
            "current_project": "Proj", "current_billing_rate": 5000,
            "billing_currency": "INR", "salary": 800000, "notes": "Good",
        }

    @pytest.fixture
    def sample_ob(self):
        return {
            "blood_group": "O+", "personal_email": "a@p.com",
            "referred_by": "HR", "bgv_status": "Done", "bgv_agency": "Co",
            "laptop_serial": "SN1", "laptop_make_model": "Dell",
            "access_card_number": "AC1", "email_id_created": "a@corp.com",
            "bank_details": {
                "account_holder_name": "Alice",
                "account_number": "12345", "ifsc_code": "HDFC001",
                "bank_name": "HDFC", "branch": "Hyd", "account_type": "Savings",
            },
            "emergency_contact": {
                "name": "Bob", "relationship": "Spouse",
                "phone": "8888",
            },
        }

    def test_returns_dict(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert isinstance(row, dict)

    def test_emp_id_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Emp ID"] == "E001"

    def test_full_name_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Full Name"] == "Alice"

    def test_date_of_joining_trimmed_to_10(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Date of Joining"] == "2023-01-01"

    def test_bank_name_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Bank Name"] == "HDFC"

    def test_ifsc_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["IFSC Code"] == "HDFC001"

    def test_emergency_contact_name_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Emergency Contact Name"] == "Bob"

    def test_blood_group_mapped(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        assert row["Blood Group"] == "O+"

    def test_empty_ob_returns_empty_strings(self, sample_emp):
        row = self._fn()(sample_emp, {})
        assert row["Bank Name"] == ""
        assert row["Emergency Contact Name"] == ""

    def test_missing_date_of_joining_returns_empty(self, sample_ob):
        emp = {"emp_id": "E002", "name": "Bob"}
        row = self._fn()(emp, sample_ob)
        assert row["Date of Joining"] == ""

    def test_all_expected_keys_present(self, sample_emp, sample_ob):
        row = self._fn()(sample_emp, sample_ob)
        for key in ("Emp ID", "Full Name", "Email", "Phone", "Designation",
                    "Department", "Status", "Annual Salary", "Blood Group",
                    "Bank Name", "IFSC Code", "Emergency Contact Name",
                    "Emergency Contact Phone", "Notes"):
            assert key in row, f"Missing key: {key}"