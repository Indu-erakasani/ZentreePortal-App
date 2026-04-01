import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, jsonify, make_response, send_file
from werkzeug.utils import secure_filename
from flask_jwt_extended import jwt_required
from bson import ObjectId
from bson.errors import InvalidId
from datetime import datetime
from extensions import mongo
from models.Employee_model import serialize_employee
from models.Onboarding_model import serialize_onboarding

export_bp = Blueprint("export", __name__)

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri")


def _flat_row(emp: dict, ob: dict) -> dict:
    bd = ob.get("bank_details", {})
    ec = ob.get("emergency_contact", {})
    doj = emp.get("date_of_joining", "")
    return {
        "Emp ID":                        emp.get("emp_id", ""),
        "Full Name":                     emp.get("name", ""),
        "Email":                         emp.get("email", ""),
        "Phone":                         emp.get("phone", ""),
        "Designation":                   emp.get("designation", ""),
        "Department":                    emp.get("department", ""),
        "Employment Type":               emp.get("employment_type", ""),
        "Date of Joining":               doj[:10] if doj else "",
        "Experience (yrs)":              emp.get("experience", ""),
        "Location":                      emp.get("location", ""),
        "Reporting Manager":             emp.get("reporting_manager", ""),
        "Status":                        emp.get("status", ""),
        "Skills":                        emp.get("skills", ""),
        "Current Client":                emp.get("current_client", ""),
        "Current Project":               emp.get("current_project", ""),
        "Billing Rate":                  emp.get("current_billing_rate", ""),
        "Billing Currency":              emp.get("billing_currency", ""),
        "Annual Salary":                 emp.get("salary", ""),
        "Blood Group":                   ob.get("blood_group", ""),
        "Personal Email":                ob.get("personal_email", ""),
        "Referred By":                   ob.get("referred_by", ""),
        "BGV Status":                    ob.get("bgv_status", ""),
        "BGV Agency":                    ob.get("bgv_agency", ""),
        "Laptop Serial":                 ob.get("laptop_serial", ""),
        "Laptop Model":                  ob.get("laptop_make_model", ""),
        "Access Card Number":            ob.get("access_card_number", ""),
        "Corporate Email":               ob.get("email_id_created", ""),
        "Bank Account Holder":           bd.get("account_holder_name", ""),
        "Bank Account Number":           bd.get("account_number", ""),
        "IFSC Code":                     bd.get("ifsc_code", ""),
        "Bank Name":                     bd.get("bank_name", ""),
        "Bank Branch":                   bd.get("branch", ""),
        "Account Type":                  bd.get("account_type", ""),
        "Emergency Contact Name":        ec.get("name", ""),
        "Emergency Contact Relation":    ec.get("relationship", ""),
        "Emergency Contact Phone":       ec.get("phone", ""),
        "Notes":                         emp.get("notes", ""),
    }


def _load_all():
    employees   = list(mongo.db.employees.find({}))
    onboardings = {o["employee_id"]: o for o in mongo.db.onboarding.find({})}
    rows = []
    for emp in employees:
        e  = serialize_employee(emp)
        ob = serialize_onboarding(onboardings.get(str(emp["_id"]), {}))
        rows.append((e, ob))
    return rows


# ── CSV export ────────────────────────────────────────────────────────────────
@export_bp.route("/employees/csv", methods=["GET"])
@jwt_required()
def export_csv():
    rows    = _load_all()
    if not rows:
        return jsonify(success=False, message="No employees to export"), 404
    output  = io.StringIO()
    flat    = [_flat_row(e, ob) for e, ob in rows]
    writer  = csv.DictWriter(output, fieldnames=flat[0].keys())
    writer.writeheader()
    writer.writerows(flat)
    resp = make_response(output.getvalue())
    resp.headers["Content-Disposition"] = (
        f"attachment; filename=employees_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    )
    resp.headers["Content-Type"] = "text/csv"
    return resp


# ── Excel export (3 sheets: Summary | Onboarding Checklist | Documents) ──────
@export_bp.route("/employees/excel", methods=["GET"])
@jwt_required()
def export_excel():
    rows = _load_all()
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Employee Summary ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Employee Summary"
    flat = [_flat_row(e, ob) for e, ob in rows]
    if flat:
        headers = list(flat[0].keys())
        for col, h in enumerate(headers, 1):
            cell       = ws.cell(row=1, column=col, value=h)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
        for r, row_data in enumerate(flat, 2):
            for c, val in enumerate(row_data.values(), 1):
                ws.cell(row=r, column=c, value=str(val) if val is not None else "")
    ws.freeze_panes = "A2"

    # ── Sheet 2: Onboarding Checklist ────────────────────────────────────────
    ws2 = wb.create_sheet("Onboarding Checklist")
    headers2 = ["Emp ID", "Name", "Checklist Item", "Status", "Remarks"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws2.column_dimensions["C"].width = 45
    for emp, ob in rows:
        for item in ob.get("checklist", []):
            ws2.append([
                emp["emp_id"], emp["name"],
                item.get("label", ""),
                "✓ Done" if item.get("done") else "Pending",
                item.get("remarks", ""),
            ])
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Documents ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Documents")
    headers3 = ["Emp ID", "Name", "Document", "Category", "Status", "Remarks"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws3.column_dimensions["C"].width = 40
    for emp, ob in rows:
        for doc in ob.get("documents", []):
            ws3.append([
                emp["emp_id"], emp["name"],
                doc.get("name", ""), doc.get("category", ""),
                doc.get("status", ""), doc.get("remarks", ""),
            ])
    ws3.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"employees_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )


# ── Single employee full profile (JSON) ──────────────────────────────────────
@export_bp.route("/employee/<emp_id>", methods=["GET"])
@jwt_required()
def export_single(emp_id):
    try:
        oid = ObjectId(emp_id)
    except InvalidId:
        return jsonify(success=False, message="Invalid ID"), 400
    emp = mongo.db.employees.find_one({"_id": oid})
    if not emp:
        return jsonify(success=False, message="Employee not found"), 404
    e  = serialize_employee(emp)
    ob = mongo.db.onboarding.find_one({"employee_id": emp_id})
    e["onboarding"] = serialize_onboarding(ob) if ob else {}
    return jsonify(success=True, data=e), 200





# ── GET /api/export/employee/<emp_id>/excel — full profile export ─────────────
@export_bp.route("/employee/<emp_id>/excel", methods=["GET"])
@jwt_required()
def export_single_excel(emp_id):
    try:
        oid = ObjectId(emp_id)
    except InvalidId:
        return jsonify(success=False, message="Invalid ID"), 400

    emp = mongo.db.employees.find_one({"_id": oid})
    if not emp:
        return jsonify(success=False, message="Employee not found"), 404

    e  = serialize_employee(emp)
    ob = mongo.db.onboarding.find_one({"employee_id": emp_id})
    ob = serialize_onboarding(ob) if ob else {}

    bd = ob.get("bank_details",      {}) or {}
    ec = ob.get("emergency_contact", {}) or {}

    wb = openpyxl.Workbook()

    # ── Helper styles ──────────────────────────────────────────────────────────
    HEAD_FILL  = PatternFill("solid", fgColor="1E40AF")
    HEAD_FONT  = Font(color="FFFFFF", bold=True, name="Calibri")
    LABEL_FONT = Font(bold=True, name="Calibri")
    VAL_FONT   = Font(name="Calibri")
    SEC_FILL   = PatternFill("solid", fgColor="EFF6FF")
    SEC_FONT   = Font(bold=True, color="1E40AF", name="Calibri")

    def write_header(ws, title, cols=2):
        ws.merge_cells(f"A1:{chr(64+cols)}1")
        cell = ws["A1"]
        cell.value         = title
        cell.font          = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        cell.fill          = HEAD_FILL
        cell.alignment     = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    def write_section(ws, row, label, cols=2):
        ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
        cell = ws[f"A{row}"]
        cell.value, cell.font, cell.fill = label, SEC_FONT, SEC_FILL
        ws.row_dimensions[row].height = 20
        return row + 1

    def write_kv(ws, row, key, value):
        kc = ws.cell(row=row, column=1, value=key);   kc.font = LABEL_FONT
        vc = ws.cell(row=row, column=2, value=str(value) if value else "—"); vc.font = VAL_FONT
        ws.row_dimensions[row].height = 18
        return row + 1

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 1 — Profile
    # ──────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Profile"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40
    write_header(ws1, f"Employee Profile — {e.get('name','')}")
    r = 2

    r = write_section(ws1, r, "Basic Information")
    for k, v in [
        ("Employee ID",      e.get("emp_id")),
        ("Full Name",        e.get("name")),
        ("Work Email",       e.get("email")),
        ("Phone",            e.get("phone")),
        ("Location",         e.get("location")),
        ("Blood Group",      ob.get("blood_group")),
        ("Personal Email",   ob.get("personal_email")),
        ("Referred By",      ob.get("referred_by")),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "Employment Details")
    for k, v in [
        ("Designation",        e.get("designation")),
        ("Department",         e.get("department")),
        ("Employment Type",    e.get("employment_type")),
        ("Status",             e.get("status")),
        ("Date of Joining",    (e.get("date_of_joining") or "")[:10]),
        ("Experience (yrs)",   e.get("experience")),
        ("Reporting Manager",  e.get("reporting_manager")),
        ("Skills",             e.get("skills")),
        ("Annual Salary",      e.get("salary")),
        ("Probation End Date", (ob.get("probation_end_date") or "")[:10]),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "Current Deployment")
    for k, v in [
        ("Current Client",   e.get("current_client")),
        ("Current Project",  e.get("current_project")),
        ("Billing Rate",     e.get("current_billing_rate")),
        ("Billing Currency", e.get("billing_currency")),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "Bank Details")
    for k, v in [
        ("Account Holder",  bd.get("account_holder_name")),
        ("Account Number",  bd.get("account_number")),
        ("IFSC Code",       bd.get("ifsc_code")),
        ("Bank Name",       bd.get("bank_name")),
        ("Branch",          bd.get("branch")),
        ("Account Type",    bd.get("account_type")),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "Emergency Contact")
    for k, v in [
        ("Name",         ec.get("name")),
        ("Relationship", ec.get("relationship")),
        ("Phone",        ec.get("phone")),
        ("Email",        ec.get("email")),
        ("Address",      ec.get("address")),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "IT Assets & Access")
    for k, v in [
        ("Laptop Serial",    ob.get("laptop_serial")),
        ("Laptop Model",     ob.get("laptop_make_model")),
        ("Access Card No.",  ob.get("access_card_number")),
        ("Corporate Email",  ob.get("email_id_created")),
    ]:
        r = write_kv(ws1, r, k, v)

    r = write_section(ws1, r, "Background Verification")
    for k, v in [
        ("BGV Status",  ob.get("bgv_status")),
        ("BGV Agency",  ob.get("bgv_agency")),
        ("BGV Remarks", ob.get("bgv_remarks")),
    ]:
        r = write_kv(ws1, r, k, v)

    if e.get("notes"):
        r = write_section(ws1, r, "Notes")
        r = write_kv(ws1, r, "Notes", e.get("notes"))

    if ob.get("hr_notes") or ob.get("it_notes"):
        r = write_section(ws1, r, "Internal Notes")
        if ob.get("hr_notes"): r = write_kv(ws1, r, "HR Notes", ob.get("hr_notes"))
        if ob.get("it_notes"): r = write_kv(ws1, r, "IT Notes", ob.get("it_notes"))

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 2 — Onboarding Checklist
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Onboarding Checklist")
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 18
    write_header(ws2, f"Onboarding Checklist — {e.get('name','')}", cols=4)

    for col, h in enumerate(["Checklist Item", "Status", "Remarks", "Updated On"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.font = Font(bold=True, name="Calibri", color="1E40AF")

    checklist  = ob.get("checklist", [])
    done_count = sum(1 for i in checklist if i.get("done"))
    for i, item in enumerate(checklist, 3):
        done = item.get("done", False)
        ws2.cell(row=i, column=1, value=item.get("label", "")).font        = VAL_FONT
        ws2.cell(row=i, column=2, value="✓ Done" if done else "Pending").font = Font(
            name="Calibri", color="15803D" if done else "C2410C", bold=done)
        ws2.cell(row=i, column=3, value=item.get("remarks", "")).font      = VAL_FONT
        ws2.cell(row=i, column=4, value=(item.get("updated_at") or "")[:10]).font = VAL_FONT

    # Summary row
    summary_row = len(checklist) + 3
    ws2.merge_cells(f"A{summary_row}:D{summary_row}")
    sc = ws2[f"A{summary_row}"]
    sc.value = f"Completion: {done_count}/{len(checklist)} items done  ({round(done_count/len(checklist)*100) if checklist else 0}%)"
    sc.font  = Font(bold=True, name="Calibri", color="15803D")
    sc.fill  = PatternFill("solid", fgColor="F0FDF4")
    ws2.freeze_panes = "A3"

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 3 — Documents
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Documents")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 30
    ws3.column_dimensions["E"].width = 22
    ws3.column_dimensions["F"].width = 18
    write_header(ws3, f"Documents — {e.get('name','')}", cols=6)

    for col, h in enumerate(["Document Name", "Category", "Status", "Remarks", "File Uploaded", "Last Updated"], 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.font = Font(bold=True, name="Calibri", color="1E40AF")

    docs = ob.get("documents", [])
    for i, doc in enumerate(docs, 3):
        ws3.cell(row=i, column=1, value=doc.get("name", "")).font     = VAL_FONT
        ws3.cell(row=i, column=2, value=doc.get("category", "")).font = VAL_FONT
        status_cell = ws3.cell(row=i, column=3, value=doc.get("status", ""))
        status_cell.font = Font(name="Calibri",
            color={"Verified": "15803D", "Pending": "C2410C",
                   "Received": "0369A1", "Waived": "64748B"}.get(doc.get("status", ""), "000000"))
        ws3.cell(row=i, column=4, value=doc.get("remarks", "")).font    = VAL_FONT
        ws3.cell(row=i, column=5, value=doc.get("file_name") or "Not uploaded").font = Font(
            name="Calibri", italic=not doc.get("file_name"))
        ws3.cell(row=i, column=6, value=(doc.get("updated_at") or "")[:10]).font = VAL_FONT

    if not docs:
        ws3.merge_cells("A3:F3")
        ws3["A3"].value = "No documents added yet"
        ws3["A3"].font  = Font(italic=True, name="Calibri", color="94A3B8")
    ws3.freeze_panes = "A3"

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 4 — Client History
    # ──────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Client History")
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["C"].width = 20
    ws4.column_dimensions["D"].width = 14
    ws4.column_dimensions["E"].width = 14
    ws4.column_dimensions["F"].width = 16
    ws4.column_dimensions["G"].width = 12
    ws4.column_dimensions["H"].width = 20
    ws4.column_dimensions["I"].width = 30
    write_header(ws4, f"Client Engagement History — {e.get('name','')}", cols=9)

    for col, h in enumerate(["Client", "Project", "Role", "Start Date", "End Date",
                              "Billing Rate", "Currency", "Work Location", "Technology"], 1):
        cell = ws4.cell(row=2, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.font = Font(bold=True, name="Calibri", color="1E40AF")

    history = e.get("client_history", [])
    for i, eng in enumerate(reversed(history), 3):
        is_active = not eng.get("end_date")
        row_fill  = PatternFill("solid", fgColor="F0FDF4") if is_active else None
        for col, val in enumerate([
            eng.get("client_name", ""),
            eng.get("project_name", ""),
            eng.get("role", ""),
            (eng.get("start_date") or "")[:10],
            (eng.get("end_date") or "Present")[:10] if eng.get("end_date") else "Present",
            eng.get("billing_rate", ""),
            eng.get("billing_currency", ""),
            eng.get("work_location", ""),
            eng.get("technology", ""),
        ], 1):
            cell = ws4.cell(row=i, column=col, value=str(val) if val else "—")
            cell.font = Font(name="Calibri",
                color="15803D" if is_active else "0F172A",
                bold=(col == 1 and is_active))
            if row_fill: cell.fill = row_fill

    if not history:
        ws4.merge_cells("A3:I3")
        ws4["A3"].value = "No client engagements recorded"
        ws4["A3"].font  = Font(italic=True, name="Calibri", color="94A3B8")
    ws4.freeze_panes = "A3"

# ── Stream ────────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Build a safe filename without werkzeug if needed
    raw_name  = f"{e.get('emp_id', 'EMP')}_{e.get('name', 'employee').replace(' ', '_')}"
    safe_name = secure_filename(raw_name)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{safe_name}_profile_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )